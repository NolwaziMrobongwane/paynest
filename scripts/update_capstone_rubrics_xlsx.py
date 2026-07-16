#!/usr/bin/env python3
"""Rewrite docs/Capstones (2).xlsx rubric sheets to match docs/assessments/*.md rubrics."""

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "Capstones (2).xlsx"
RUBRICS_DIR = ROOT / "docs" / "capestone-rubrics"

HEAD_ROW1 = ("Max band pts", 5, 10, 20, 25)
HEAD_ROW2 = ("Criteria", "1 - Needs work", "2 - Getting there", "3 - Acceptable", "4 - Outstanding")

Rows = list[tuple[str, str, str, str, str]]

C1: Rows = [
    (
        "Architecture & domain modelling (25%)",
        "No clear domain boundaries; god classes; OrderService API unclear or missing.",
        "Classes exist but responsibilities overlap; extending fields risks breaking callers.",
        "Clear Product, Customer, Order, OrderItem split; coherent OrderService; stable callers.",
        "Level 3 plus exemplary separation and a credible story for adding catalogue fields without rewriting checkout.",
    ),
    (
        "Correctness & business rules (30%)",
        "Line totals or grand total wrong/missing; quantities unchecked; summary disagrees with internals.",
        "Mostly correct totals with occasional inconsistencies or weak validation.",
        "Subtotals and grand total match stated rules; quantities validated; summary matches internal computation.",
        "Level 3 plus strong edge-case handling (empty order, multiples) or explicit fail-fast with clear messages.",
    ),
    (
        "Testing & verification (15%)",
        "No meaningful tests and/or mvn test fails on submission.",
        "Sparse tests; gaps on non-trivial arithmetic or collection behaviour.",
        "Tests cover totals and non-trivial cases; mvn test green; regressions would be caught.",
        "Level 3 plus tests that would catch subtle collection/total regressions and document intent.",
    ),
    (
        "Code quality & maintainability (20%)",
        "Hard to read; inconsistent style; public fields; dead experimental paths in main.",
        "Readable in places; inconsistent encapsulation or naming.",
        "Readable, consistent style; sensible encapsulation; clean demo entry point.",
        "Level 3 plus professional structure, naming, and maintainability across packages.",
    ),
    (
        "Documentation & communication (10%)",
        "No setup/run guidance for reviewers.",
        "Minimal notes; hard to reproduce the demo.",
        "Setup/run instructions; comments where rules are non-obvious; optional diagram.",
        "Level 3 plus a precise diagram or short note tying design choices to future change.",
    ),
]

C2: Rows = [
    (
        "Architecture & polymorphism (25%)",
        "No credible PaymentMethod abstraction; checkout riddled with type switches.",
        "Interface present but callers still depend on concrete types or large branching.",
        "Clean contract; checkout depends on abstraction; minimal branching at call sites.",
        "Level 3 plus elegant extensibility: new rail needs new class + wiring, not Order edits.",
    ),
    (
        "Correctness & business rules (30%)",
        "Wrong amount sent to payment; rail labels wrong; success path inconsistent with order total.",
        "Mostly correct with occasional mismatches or unclear success semantics.",
        "Order total flows into payment; rail labels coherent; success path matches domain expectations.",
        "Level 3 plus robust handling of failure/edge paths aligned with stated contract.",
    ),
    (
        "Testing & verification (20%)",
        "No tests proving polymorphism; mvn test fails.",
        "Limited tests; polymorphism not actually exercised in tests.",
        "Tests exercise at least two payment implementations; totals/checkout regression covered.",
        "Level 3 plus strong coverage of checkout integration and extension safety.",
    ),
    (
        "Code quality & maintainability (15%)",
        "Duplicated total logic; messy rail classes; tangled dependencies.",
        "Some duplication or uneven rail implementations.",
        "Readable per-rail code; no duplicated total calculation paths.",
        "Level 3 plus consistent patterns across rails and clear package boundaries.",
    ),
    (
        "Documentation & communication (10%)",
        "No rationale for design; run steps unclear.",
        "Brief notes only; diagram missing or misleading.",
        "Short rationale (why interfaces vs mega-method); clear run instructions; optional accurate diagram.",
        "Level 3 plus crisp diagram and explicit answer to 'what changes for a new rail?'",
    ),
]

C3: Rows = [
    (
        "Architecture & routing design (25%)",
        "Tightly coupled routing; no clear ordering story; hard to add rules/providers.",
        "Some layering; ordering or config story incomplete.",
        "Clear rules/engine/providers split; documented rule ordering; extensible RoutingConfig usage.",
        "Level 3 plus policy clarity under change (new provider/rule) with minimal blast radius.",
    ),
    (
        "Correctness & business rules (30%)",
        "Decisions wrong or arbitrary; fallback dishonest; risk not derivable from inputs.",
        "Partial correctness; edge cases weak; risk mapping thin.",
        "Decisions match declared predicates; honest fallback semantics; justified RiskLevel mapping.",
        "Level 3 plus comprehensive edge coverage and no silent default-first-provider behaviour.",
    ),
    (
        "Testing & verification (20%)",
        "No tests for priority/tie-break/fallback; mvn test fails.",
        "Sparse tests; key routing paths untested.",
        "Tests encode priority, tie-break, and fallback; regressions fail loudly.",
        "Level 3 plus risk/rule matrix tests and regression harness for config changes.",
    ),
    (
        "Auditability & operational realism (15%)",
        "No traceable reasons; applied rules list unused; logs unusable under demo load.",
        "Minimal audit trail; reasons vague.",
        "RouteDecision reasons and applied-rules lists trustworthy; DecisionLogger readable.",
        "Level 3 plus structured audit output suitable for support replay.",
    ),
    (
        "Documentation & communication (10%)",
        "No design note; scenarios not reproducible from submission.",
        "Partial note; missing decision traces.",
        "Design note with priority diagram/timeline; two example decision traces.",
        "Level 3 plus instructor can replay scenarios from documentation alone.",
    ),
]

C4: Rows = [
    (
        "Architecture & reliability design (25%)",
        "No persistence boundary; reporting hard-coded; mocks in the production path.",
        "Some abstraction but store/registry split unclear or reporting disconnected.",
        "Clear pipeline stages; persistence abstraction; reporting reads real store data.",
        "Level 3 plus coherent layering for future HA (even if single-node today).",
    ),
    (
        "Correctness & business rules (30%)",
        "Idempotency broken; illegal status transitions; failures corrupt state silently.",
        "Partial idempotency; status model shaky; failure behaviour undocumented.",
        "Idempotency proven; coherent statuses; failure behaviour documented and tested.",
        "Level 3 plus transactional boundaries or concurrency story where justified.",
    ),
    (
        "Testing & verification (20%)",
        "No duplicate-key or failure tests; mvn test fails.",
        "Limited tests; reporting not validated on fixtures.",
        "Duplicate-key tests; failure injections; reporting validated on fixtures.",
        "Level 3 plus broad regression suite for persistence edge cases.",
    ),
    (
        "Operational realism (15%)",
        "No DDL discipline; data dir chaos; reports not reproducible from clean DB.",
        "Basic persistence; weak runbook or hygiene.",
        "DDL/schema artefact; data/ hygiene; honest aggregates from persisted rows.",
        "Level 3 plus runbook for wipe/rebuild and incident narrative quality.",
    ),
    (
        "Documentation & communication (10%)",
        "Missing schema/runbook; cannot reproduce reports.",
        "Thin notes on idempotency or schema.",
        "Incident narrative + schema/runbook; instructor can reproduce reports locally.",
        "Level 3 plus crisp ops story for duplicate POST storm and recovery.",
    ),
]

C5: Rows = [
    (
        "Architecture & hybrid design (25%)",
        "Monolithic glue; no queue/worker story; merge policy undocumented; audit missing.",
        "Partial separation; weak audit or unclear merge policy.",
        "Queue, evaluation, persistence separated; documented merge policy; AiDecisionStore design clear.",
        "Level 3 plus clean extension path for new signals without rewriting hybrid core.",
    ),
    (
        "Correctness & resilience (30%)",
        "Crashes on bad LLM output/offline Ollama; state inconsistent with TransactionRecord.",
        "Some try/catch but silent wrong states or hangs risk.",
        "Offline degradation; malformed AI handled; TransactionRecord fields stay consistent with policy.",
        "Level 3 plus timeouts/retries where appropriate; no hangs on slow HTTP.",
    ),
    (
        "Testing & verification (20%)",
        "Tests rely on live model creativity; mvn test flaky or absent.",
        "Limited parser/hybrid tests.",
        "Parser tests; hybrid offline parity; monitoring smoke without flaky LLM asserts.",
        "Level 3 plus fixture-based LLM payloads and strong regression coverage.",
    ),
    (
        "Operational realism (15%)",
        "Unbounded queue growth; no Ollama ops notes; thread policy absent.",
        "Basic worker; weak back-pressure or ops story.",
        "Thread/queue policy stated; Ollama ops notes; avoids unbounded growth without strategy.",
        "Level 3 plus thoughtful shutdown/back-pressure and lab limitations honestly stated.",
    ),
    (
        "Documentation & communication (10%)",
        "No ops note or demo checklist.",
        "Partial ops note; checklist incomplete.",
        "Ops note (threading, bounds, Ollama down) + demo checklist; optional accurate sequence diagram.",
        "Level 3 plus diagrams and explicit limitation discussion per assessment brief.",
    ),
]

# Order matches workbook tabs. candidates: try first name that exists (supports re-run after rename).
SHEETS: list[dict] = [
    {
        "candidates": ["Capstone 1 - Core Commerce"],
        "new_title": "Capstone 1 - Core Commerce",
        "slug": "capstone-01",
        "rows": C1,
    },
    {
        "candidates": ["Capstone 2 - OOP Payments"],
        "new_title": "Capstone 2 - OOP Payments",
        "slug": "capstone-02",
        "rows": C2,
    },
    {
        "candidates": [
            "Capstone 3 - Routing & Risk",
            "Capestone 3 - Adaptive Payment ",
        ],
        "new_title": "Capstone 3 - Routing & Risk",
        "slug": "capstone-03",
        "rows": C3,
    },
    {
        "candidates": [
            "Capstone 4 - Persistence & Ops",
            "Capestone 4 - Persistence, Stat",
        ],
        "new_title": "Capstone 4 - Persistence & Ops",
        "slug": "capstone-04",
        "rows": C4,
    },
    {
        "candidates": [
            "Capstone 5 - Monitoring & AI",
            "Capstone 5 - Real-Time Agentic ",
        ],
        "new_title": "Capstone 5 - Monitoring & AI",
        "slug": "capstone-05",
        "rows": C5,
    },
]


def apply_rubric(ws, rows: Rows) -> None:
    for c, val in enumerate(HEAD_ROW1, start=1):
        ws.cell(row=1, column=c, value=val)
    for c, val in enumerate(HEAD_ROW2, start=1):
        ws.cell(row=2, column=c, value=val)
    for i, row in enumerate(rows, start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=val)
    for r in range(3 + len(rows), 40):
        for c in range(1, 6):
            ws.cell(row=r, column=c).value = None


def resolve_worksheet(wb: openpyxl.Workbook, candidates: list[str]):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    raise SystemExit(f"No sheet found among {candidates!r}; have {wb.sheetnames!r}")


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    for cfg in SHEETS:
        ws = resolve_worksheet(wb, cfg["candidates"])
        apply_rubric(ws, cfg["rows"])
        ws.title = str(cfg["new_title"])[:31]

    wb.save(XLSX)
    print("Updated", XLSX)

    RUBRICS_DIR.mkdir(parents=True, exist_ok=True)
    (RUBRICS_DIR / "README.md").write_text(
        """# Capestone rubrics (mirror)

This folder holds **CSV mirrors** of the rubrics aligned with [docs/assessments/](../assessments/README.md).

- Full project briefs: [docs/assessments/README.md](../assessments/README.md)
- Excel workbook for scoring: [Capstones (2).xlsx](../Capstones%20(2).xlsx)

Each `capstone-0N-rubric.csv` has columns: Criterion, Level 1–4 descriptors (Needs work → Outstanding).
""",
        encoding="utf-8",
    )

    for cfg in SHEETS:
        lines = [
            "Criterion,Level_1_Needs_work,Level_2_Getting_there,Level_3_Acceptable,Level_4_Outstanding"
        ]
        for row in cfg["rows"]:
            esc = [str(x).replace('"', '""') for x in row]
            lines.append(",".join(f'"{p}"' for p in esc))
        (RUBRICS_DIR / f"{cfg['slug']}-rubric.csv").write_text("\n".join(lines) + "\n", encoding="utf-8")

    print("Wrote CSV mirrors to", RUBRICS_DIR)


if __name__ == "__main__":
    main()
